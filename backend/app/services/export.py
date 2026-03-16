"""Export helpers — Banana TSV, styled Excel, semicolon CSV."""
from __future__ import annotations
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def fmt_swiss(val) -> str:
    """Format a number in Swiss style: 1'234.56"""
    if val is None or val == "" or (isinstance(val, float) and pd.isna(val)):
        return ""
    num = float(val)
    negative = num < 0
    if negative:
        num = abs(num)
    integer_part = int(num)
    decimal_part = f"{num - integer_part:.2f}"[1:]
    int_str = f"{integer_part:,}".replace(",", "\u2019")
    result = f"{int_str}{decimal_part}"
    if negative:
        result = f"-{result}"
    return result


def df_to_styled_excel(df: pd.DataFrame) -> bytes:
    """Export DataFrame to a formatted .xlsx with headers, borders, colors."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Buchhaltung"

    headers = list(df.columns)
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    col_widths = {
        "Nr": 6, "Datum": 12, "Beleg": 10, "Rechnung": 10,
        "Beschreibung": 48, "KtSoll": 8, "KtHaben": 8,
        "Betrag CHF": 14, "MwStUSt-Code": 16, "Art Betrag": 10,
        "MwSt-%": 9, "Gebuchte MwStUSt CHF": 20, "KS3": 6,
    }

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(
            header, 12
        )

    red_font = Font(color="FF0000", size=10)
    normal_font = Font(size=10)
    number_cols = {"Betrag CHF", "Gebuchte MwStUSt CHF", "MwSt-%"}

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, header in enumerate(headers, 1):
            val = row[header]
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

            if header in number_cols and val != "" and val is not None:
                try:
                    numval = float(val)
                    cell.value = numval
                    cell.number_format = (
                        "#,##0.00" if header != "MwSt-%" else "0.00"
                    )
                    cell.alignment = Alignment(horizontal="right")
                    cell.font = red_font if numval < 0 else normal_font
                except (ValueError, TypeError):
                    cell.value = val
                    cell.font = normal_font
            elif header in ("KtSoll", "KtHaben", "Nr"):
                cell.value = val
                cell.alignment = Alignment(horizontal="center")
                cell.font = normal_font
            else:
                cell.value = val if val != "" else None
                cell.font = normal_font

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def df_to_banana_tsv(df: pd.DataFrame) -> str:
    """Export DataFrame to Banana Accounting tab-separated import format."""
    banana_cols = [
        "Date", "Description", "AccountDebit", "AccountCredit", "Amount", "VatCode",
    ]
    lines = ["\t".join(banana_cols)]

    for _, row in df.iterrows():
        datum = str(row.get("Datum", ""))
        if datum and "." in datum:
            parts = datum.split(".")
            if len(parts) == 3:
                datum = f"{parts[2]}-{parts[1]}-{parts[0]}"

        description = str(row.get("Beschreibung", ""))
        ktsoll = str(row.get("KtSoll", ""))
        kthaben = str(row.get("KtHaben", ""))
        betrag = row.get("Betrag CHF", 0)

        try:
            amount = f"{float(betrag):.2f}" if betrag else ""
        except (ValueError, TypeError):
            amount = ""

        vatcode = str(row.get("MwStUSt-Code", ""))
        fields = [datum, description, ktsoll, kthaben, amount, vatcode]
        lines.append("\t".join(fields))

    return "\n".join(lines)


def df_to_csv(df: pd.DataFrame) -> str:
    """Export DataFrame to semicolon-separated CSV."""
    buf = io.StringIO()
    df.to_csv(buf, index=False, sep=";")
    return buf.getvalue()
